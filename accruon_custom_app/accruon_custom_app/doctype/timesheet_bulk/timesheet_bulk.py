# Copyright (c) 2024, Josin Benny and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import openpyxl


class TimesheetBulk(Document):
	pass

def read_excel():
	wb = openpyxl.load_workbook(get_absolute_path('/files/10-EMDAD-HABSHAN-OCT-2024 (2).xlsx'))
	ws = wb.active
	emps = frappe.get_all("Employee",{"Status":"active","name":"UI102"},pluck="name")
	ts = []

	print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
	for i in range(1,ws.max_row+1):
		cell = f"B{i}"
		if ws[cell].value in emps:
			print("yeyeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee")
		else:
			print("yoyo")
	# print('The value in cell A1 is: '+ws['C6'].value)
		

def get_absolute_path(file_name, is_private=False):
	if(file_name.startswith('/files/')):
		file_name = file_name[7:]
	return frappe.utils.get_bench_path()+ "/sites/" + frappe.utils.get_path('private' if is_private else 'public', 'files', file_name)[2:]